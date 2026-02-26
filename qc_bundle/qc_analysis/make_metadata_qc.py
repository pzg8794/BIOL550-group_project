#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
import zipfile
from collections import OrderedDict
from copy import copy
from dataclasses import dataclass
from typing import Iterable

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class QcResult:
    overall: str  # PASS | WARN | FAIL | MISSING
    fail_count: int = 0
    warn_count: int = 0
    pass_count: int = 0


def _copy_cell_style(src: openpyxl.cell.cell.Cell, dst: openpyxl.cell.cell.Cell) -> None:
    dst.font = copy(src.font)
    dst.border = copy(src.border)
    dst.fill = copy(src.fill)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)
    dst.alignment = copy(src.alignment)


def _copy_sheet_dimensions(
    ws_in: openpyxl.worksheet.worksheet.Worksheet,
    ws_out: openpyxl.worksheet.worksheet.Worksheet,
) -> None:
    ws_out.sheet_format = copy(ws_in.sheet_format)
    ws_out.sheet_properties = copy(ws_in.sheet_properties)
    ws_out.merged_cells = copy(ws_in.merged_cells)
    ws_out.page_margins = copy(ws_in.page_margins)
    ws_out.page_setup = copy(ws_in.page_setup)
    ws_out.print_options = copy(ws_in.print_options)
    ws_out.views = copy(ws_in.views)
    ws_out.freeze_panes = ws_in.freeze_panes

    for col_letter, dim in ws_in.column_dimensions.items():
        ws_out.column_dimensions[col_letter].width = dim.width
        ws_out.column_dimensions[col_letter].hidden = dim.hidden

    for row_idx, dim in ws_in.row_dimensions.items():
        ws_out.row_dimensions[row_idx].height = dim.height
        ws_out.row_dimensions[row_idx].hidden = dim.hidden


def parse_fastqc_zip_summary(zip_path: str) -> QcResult:
    if not os.path.exists(zip_path):
        return QcResult(overall="MISSING")

    try:
        with zipfile.ZipFile(zip_path) as zf:
            summary_members = [n for n in zf.namelist() if n.endswith("/summary.txt") or n == "summary.txt"]
            if not summary_members:
                return QcResult(overall="MISSING")

            content = zf.read(summary_members[0]).decode("utf-8", errors="replace")
    except zipfile.BadZipFile:
        return QcResult(overall="MISSING")

    statuses: list[str] = []
    for line in content.splitlines():
        if not line.strip():
            continue
        status = line.split("\t", 1)[0].strip().upper()
        if status in {"PASS", "WARN", "FAIL"}:
            statuses.append(status)

    fail_count = sum(1 for s in statuses if s == "FAIL")
    warn_count = sum(1 for s in statuses if s == "WARN")
    pass_count = sum(1 for s in statuses if s == "PASS")

    if fail_count:
        overall = "FAIL"
    elif warn_count:
        overall = "WARN"
    elif statuses:
        overall = "PASS"
    else:
        overall = "MISSING"

    return QcResult(overall=overall, fail_count=fail_count, warn_count=warn_count, pass_count=pass_count)


def iter_run_groups(
    ws: openpyxl.worksheet.worksheet.Worksheet, run_col: int = 1, start_row: int = 2
) -> Iterable[tuple[str, list[int]]]:
    grouped: "OrderedDict[str, list[int]]" = OrderedDict()
    for r in range(start_row, ws.max_row + 1):
        run = ws.cell(r, run_col).value
        if run is None or str(run).strip() == "":
            continue
        run = str(run).strip()
        grouped.setdefault(run, []).append(r)
    return grouped.items()


def merge_repeated_leading_cells(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
    max_col: int,
) -> None:
    if end_row <= start_row:
        return

    first_diff_col: int | None = None
    for c in range(1, max_col + 1):
        first_val = ws.cell(start_row, c).value
        all_same = True
        for r in range(start_row + 1, end_row + 1):
            if ws.cell(r, c).value != first_val:
                all_same = False
                break
        if not all_same:
            first_diff_col = c
            break

    merge_upto = (first_diff_col - 1) if first_diff_col is not None else max_col
    if merge_upto < 1:
        return

    for c in range(1, merge_upto + 1):
        first_val = ws.cell(start_row, c).value
        if first_val is None or str(first_val) == "":
            continue

        all_same = True
        for r in range(start_row + 1, end_row + 1):
            if ws.cell(r, c).value != first_val:
                all_same = False
                break
        if not all_same:
            continue

        for r in range(start_row + 1, end_row + 1):
            ws.cell(r, c).value = None
        ws.merge_cells(start_row=start_row, start_column=c, end_row=end_row, end_column=c)


def build_output_workbook(
    input_xlsx: str,
    fastqc_dir: str,
    output_xlsx: str,
    sheet_name: str | None = None,
) -> None:
    wb_in = openpyxl.load_workbook(input_xlsx)
    ws_in = wb_in[sheet_name] if sheet_name else wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title
    _copy_sheet_dimensions(ws_in, ws_out)

    base_cols = ws_in.max_column
    extra_headers = [
        ("Read", None),
        ("FastQC QC", None),
        ("FastQC fail_count", None),
        ("FastQC warn_count", None),
        ("FastQC HTML", "Link to report"),
    ]
    out_cols = base_cols + len(extra_headers)

    # Header row (preserve styles)
    for c in range(1, base_cols + 1):
        src = ws_in.cell(1, c)
        dst = ws_out.cell(1, c, value=src.value)
        _copy_cell_style(src, dst)

    header_font = copy(ws_out.cell(1, 1).font) if ws_out.cell(1, 1).font else Font(bold=True)
    for i, (h, _) in enumerate(extra_headers, start=1):
        c = base_cols + i
        cell = ws_out.cell(1, c, value=h)
        cell.font = header_font

    # Fill styles for QC status
    fill_pass = PatternFill(patternType="solid", fgColor="FFC6EFCE")
    fill_warn = PatternFill(patternType="solid", fgColor="FFFFEB9C")
    fill_fail = PatternFill(patternType="solid", fgColor="FFFFC7CE")
    fill_missing = PatternFill(patternType="solid", fgColor="FFD9D9D9")

    qc_col_idx = base_cols + 2
    fail_count_col_idx = base_cols + 3
    warn_count_col_idx = base_cols + 4
    html_col_idx = base_cols + 5

    out_row = 2
    run_output_ranges: list[tuple[int, int]] = []
    for run, row_indices in iter_run_groups(ws_in, run_col=1, start_row=2):
        run_start_row = out_row
        for read in (1, 2):
            src_row = row_indices[min(read - 1, len(row_indices) - 1)]

            for c in range(1, base_cols + 1):
                src = ws_in.cell(src_row, c)
                dst = ws_out.cell(out_row, c, value=src.value)
                _copy_cell_style(src, dst)

            ws_out.cell(out_row, base_cols + 1, value=read)

            zip_name = f"{run}_{read}_fastqc.zip"
            zip_path = os.path.join(fastqc_dir, zip_name)
            qc = parse_fastqc_zip_summary(zip_path)
            qc_cell = ws_out.cell(out_row, qc_col_idx, value=qc.overall)
            ws_out.cell(out_row, fail_count_col_idx, value=qc.fail_count)
            ws_out.cell(out_row, warn_count_col_idx, value=qc.warn_count)

            if qc.overall == "PASS":
                qc_cell.fill = fill_pass
            elif qc.overall == "WARN":
                qc_cell.fill = fill_warn
            elif qc.overall == "FAIL":
                qc_cell.fill = fill_fail
            else:
                qc_cell.fill = fill_missing

            html_name = f"{run}_{read}_fastqc.html"
            html_path = os.path.join(fastqc_dir, html_name)
            display = f"../{html_name}" if os.path.exists(html_path) else html_name
            html_cell = ws_out.cell(out_row, html_col_idx, value=display)
            if os.path.exists(html_path):
                html_cell.hyperlink = os.path.relpath(html_path, os.path.dirname(output_xlsx))
                html_cell.style = "Hyperlink"

            out_row += 1
        run_output_ranges.append((run_start_row, out_row - 1))

    # Make sure new columns are visible with a reasonable width
    for c in range(base_cols + 1, out_cols + 1):
        col_letter = get_column_letter(c)
        if ws_out.column_dimensions[col_letter].width is None:
            ws_out.column_dimensions[col_letter].width = 18

    # Merge repeated leading cells within each (Run, Read1/Read2) block:
    # merge from column 1 up to the first column that changes.
    for start, end in run_output_ranges:
        merge_repeated_leading_cells(ws_out, start_row=start, end_row=end, max_col=out_cols)

    os.makedirs(os.path.dirname(output_xlsx), exist_ok=True)
    wb_out.save(output_xlsx)


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Split paired-end runs into 2 rows (R1/R2) and add FastQC QC status with colored cells."
        )
    )
    parser.add_argument(
        "--input-xlsx",
        default=os.path.join(os.path.dirname(__file__), "metadata.xlsx"),
        help="Path to input metadata.xlsx",
    )
    parser.add_argument(
        "--fastqc-dir",
        default=os.path.abspath(os.path.join(os.path.dirname(__file__), "..")),
        help="Directory containing *_fastqc.zip and *_fastqc.html files",
    )
    parser.add_argument(
        "--output-xlsx",
        default=os.path.join(os.path.dirname(__file__), "metadata_qc.xlsx"),
        help="Path to write output xlsx",
    )
    parser.add_argument("--sheet", default=None, help="Sheet name (default: active sheet)")
    args = parser.parse_args()

    build_output_workbook(
        input_xlsx=os.path.abspath(args.input_xlsx),
        fastqc_dir=os.path.abspath(args.fastqc_dir),
        output_xlsx=os.path.abspath(args.output_xlsx),
        sheet_name=args.sheet,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
